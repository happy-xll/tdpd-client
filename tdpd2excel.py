import click
import io
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from lib_tdpd import authorized_get, get_middata_group

# 从配置文件读取默认值
DEFAULT_GROUP = get_middata_group()

# command 入口
@click.command()
@click.option('--excel', '-e', default='tdpd_param_value_by_group.xlsx', type=str, help='请指定导出文件')
@click.option('--group', '-g', default=DEFAULT_GROUP, type=int, help=f'请指定参数名称 (默认使用配置文件middata_group={DEFAULT_GROUP})')
@click.option('--year', '-y', default=None, type=int, help='请指定数据年份, eg: 2022')
def fun(excel, group, year):
    paramgroup2excel(excel, group, year)


def paramgroup2excel(to_file, param_group, year):
    # 1. 获取参数列表， 也就是group
    url = 'v1/params?group={}'.format(param_group)
    params = authorized_get(url)
    if not to_file:
        to_file = './tdpd_pg_{}.xlsx'.format(param_group)

    # 使用 engine='openpyxl' 以便后续操作
    with pd.ExcelWriter(to_file, engine='openpyxl') as writer:
        for item in params.get('items'):
            sheet_d = load_sheet_data(item.get('pk_param'), year)
            sheet_name = item.get('param_name')
            sheet_d.to_excel(writer, sheet_name=sheet_name, index=False)

        # 获取 workbook 和 worksheets 来添加条件格式和公式列
        workbook = writer.book
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        for sheet in workbook.worksheets:
            # 跳过行数不足的工作表
            if sheet.max_row < 2:
                continue

            print(f"处理工作表: {sheet.title} (共 {sheet.max_row - 1} 行数据, {sheet.max_column} 列)")

            # 先记录所有需要添加状态列的数值列（保存原始列号）
            columns_to_add_formula = []
            skipped_columns = []

            for col in range(1, sheet.max_column + 1):
                col_letter = get_column_letter(col)
                header = sheet[f"{col_letter}1"].value

                # 跳过 year 和 month 等维度列
                if header in ['year', 'month', 'day', 'hour', 'minute', 'fk_zone', 'fk_param']:
                    continue

                # 检查该列是否有足够的数据
                if sheet.max_row < 3:
                    continue

                # 检查该列是否为数值类型（采样检查）
                has_numeric = False
                sample_values = []
                for row in range(2, min(10, sheet.max_row + 1)):  # 检查更多行
                    cell = sheet[f"{col_letter}{row}"]
                    if cell.value is not None:
                        sample_values.append(cell.value)
                        if isinstance(cell.value, (int, float)):
                            has_numeric = True
                            break

                if has_numeric:
                    columns_to_add_formula.append((col, col_letter, header))
                else:
                    skipped_columns.append((col, col_letter, header, sample_values[:3]))

            # 打印跳过的列信息
            if skipped_columns:
                print(f"  跳过的列（无数值数据）:")
                for col, col_letter, header, samples in skipped_columns:
                    print(f"    {header} ({col_letter}): 采样值={samples}")

            # 正序插入状态列，动态跟踪插入位置
            inserted_count = 0

            for col, col_letter, header in columns_to_add_formula:
                # 计算当前原始列的实际位置（前面已插入了 inserted_count 列）
                current_col_num = col + inserted_count
                current_col_letter = get_column_letter(current_col_num)

                # 在当前列后面插入一列
                new_col_num = current_col_num + 1
                new_col_letter = get_column_letter(new_col_num)
                sheet.insert_cols(new_col_num)
                inserted_count += 1  # 增加已插入列数

                # 设置新列标题
                new_header = sheet[f"{new_col_letter}1"]
                new_header.value = f"{header}_状态"

                # 获取数据行数
                last_row = sheet.max_row
                abs_range = f"${current_col_letter}$2:${current_col_letter}${last_row}"

                # 在新列中添加公式（引用当前原始列）
                for row in range(2, last_row + 1):
                    cell = sheet[f"{new_col_letter}{row}"]
                    # Excel 公式：判断是否超出 2σ
                    formula = f'=IF(ABS({current_col_letter}{row}-AVERAGE({abs_range}))>2*STDEV({abs_range}),"超出2σ","正常")'
                    cell.value = formula

                # 添加条件格式到状态列：只有"超出2σ"标黄色
                status_range = f"{new_col_letter}2:{new_col_letter}{last_row}"
                rule_status = CellIsRule(operator='equal', formula=['"超出2σ"'], stopIfTrue=True, fill=yellow_fill)
                sheet.conditional_formatting.add(status_range, rule_status)

                # 添加条件格式到原始数据列：使用完整的 ABS 公式，与 IF 公式完全一致
                data_range = f"{current_col_letter}2:{current_col_letter}{last_row}"
                # 使用 FormulaRule 来支持完整的表达式公式
                formula_abs = f'ABS({current_col_letter}2-AVERAGE({abs_range}))>2*STDEV({abs_range})'
                rule_abs = FormulaRule(formula=[formula_abs], stopIfTrue=True, fill=red_fill)
                sheet.conditional_formatting.add(data_range, rule_abs)

                print(f"  列 '{header}' (原始列{current_col_letter}, 状态列{new_col_letter}): 已插入公式列")

def load_sheet_data(param_id, year=None):
    if year:
        url = 'v1/param/{}/values?only_data=1&pivot=1&patch=1&year={}'.format(param_id, year)
    else:
        url = 'v1/param/{}/values?only_data=1&patch=1&pivot=1'.format(param_id)
    djson = authorized_get(url, 'json')
    dfn = pd.read_json(io.StringIO(djson), orient='split')
    return dfn

if __name__ == '__main__':
    fun()
    # paramgroup2excel('test3.xlsx', 1, None)
